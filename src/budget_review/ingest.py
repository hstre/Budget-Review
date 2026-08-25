"""Turn supported proposal and budget files into one auditable source text."""

from __future__ import annotations

import csv
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree


class IngestError(ValueError):
    pass


@dataclass(frozen=True)
class SourceBundle:
    document_id: str
    text: str
    source_files: tuple[str, ...]


def _read_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - depends on optional extra
        raise IngestError("PDF support requires: pip install 'budget-review[documents]'") from exc
    pages = []
    for index, page in enumerate(PdfReader(str(path)).pages, start=1):
        pages.append(f"[PAGE {index}]\n{page.extract_text() or ''}")
    return "\n\n".join(pages)


def _read_docx(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile) as exc:
        raise IngestError(f"invalid DOCX file: {path}") from exc
    root = ElementTree.fromstring(xml)
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    paragraphs: list[str] = []
    for paragraph in root.iter(namespace + "p"):
        text = "".join(node.text or "" for node in paragraph.iter(namespace + "t"))
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def _read_delimited(path: Path, delimiter: str) -> str:
    lines: list[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row_index, row in enumerate(csv.reader(handle, delimiter=delimiter), start=1):
            cells = [
                f"{_column_name(index)}{row_index}={value}" for index, value in enumerate(row, 1)
            ]
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def _column_name(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _read_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on optional extra
        raise IngestError("XLSX support requires: pip install 'budget-review[documents]'") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"[SHEET {sheet.title}]")
        for row in sheet.iter_rows():
            cells = [f"{cell.coordinate}={cell.value}" for cell in row if cell.value is not None]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def read_source(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8")
    if suffix == ".pdf":
        return _read_pdf(path)
    if suffix == ".docx":
        return _read_docx(path)
    if suffix == ".csv":
        return _read_delimited(path, ",")
    if suffix == ".tsv":
        return _read_delimited(path, "\t")
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise IngestError(f"unsupported input format: {suffix or '<none>'}")


def ingest(paths: list[Path], document_id: str | None = None) -> SourceBundle:
    if not paths:
        raise IngestError("at least one input file is required")
    for path in paths:
        if not path.is_file():
            raise IngestError(f"input file does not exist: {path}")
    texts = [read_source(path) for path in paths]
    if len(paths) == 1:
        combined = texts[0]
    else:
        combined = "\n\n".join(
            f"[[SOURCE: {path.name}]]\n{text}" for path, text in zip(paths, texts, strict=True)
        )
    if not combined.strip():
        raise IngestError("input contains no extractable text")
    inferred = re.sub(r"[^A-Za-z0-9_.-]+", "-", paths[0].stem).strip("-") or "document"
    return SourceBundle(
        document_id=document_id or inferred,
        text=combined,
        source_files=tuple(str(path) for path in paths),
    )


__all__ = ["IngestError", "SourceBundle", "ingest", "read_source"]
